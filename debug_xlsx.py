from pathlib import Path
from openpyxl import load_workbook

p = Path(r'c:\Users\User\Downloads\AgriScheme_backend\scheme_data.xlsx')
print('exists=', p.exists())
wb = load_workbook(p, data_only=True)
print('sheets=', wb.sheetnames)
for ws in wb.worksheets:
    print('---', ws.title)
    for row in list(ws.iter_rows(values_only=True, max_row=min(ws.max_row, 5))):
        print(row)
